from __future__ import annotations

from io import BytesIO
from typing import Dict, List
import re

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .config import TABLE_ALIASES


def diagnose_workbook_structure(file_bytes: bytes) -> Dict[str, List[str]]:
    """Return a lightweight workbook quality report before full parsing."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    issues: List[str] = []
    warnings: List[str] = []
    discovered_tables: List[str] = []

    for ws in wb.worksheets:
        for table in ws.tables.values():
            canonical_name = TABLE_ALIASES.get(table.name, table.name)
            discovered_tables.append(canonical_name)

            ref = ws[table.ref]
            data = [[c.value for c in row] for row in ref]
            if not data:
                issues.append(f"Table '{table.name}' in sheet '{ws.title}' is empty.")
                continue

            headers = [str(h).strip() if h is not None else "" for h in data[0]]
            if any(not h for h in headers):
                issues.append(f"Table '{table.name}' in sheet '{ws.title}' has blank header cells.")

            row_count = len(data) - 1
            if row_count == 0:
                warnings.append(f"Table '{table.name}' has headers but no data rows.")

            if len(set(headers)) != len(headers):
                warnings.append(f"Table '{table.name}' has duplicate column headers.")

    if not discovered_tables:
        issues.append("No Excel tables were detected. Inputs must be formatted as named Excel Tables, not plain cell ranges.")

    return {
        "issues": issues,
        "warnings": warnings,
        "tables": sorted(discovered_tables),
    }


def load_tables_from_excel(file_bytes: bytes) -> Dict[str, pd.DataFrame]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    tables: Dict[str, pd.DataFrame] = {}

    for ws in wb.worksheets:
        for table in ws.tables.values():
            ref = ws[table.ref]
            data = [[c.value for c in row] for row in ref]
            if not data:
                continue
            headers = [str(h) for h in data[0]]
            rows = data[1:]
            df = pd.DataFrame(rows, columns=headers)
            tname = TABLE_ALIASES.get(table.name, table.name)
            tables[tname] = df

    return tables


def _write_df_as_table(ws, df: pd.DataFrame, table_name: str):
    # Always reset worksheet content and table definitions before writing.
    # This guarantees a single header block + single table, even if this
    # helper is called repeatedly for the same worksheet object.
    for existing_table_name in list(ws.tables.keys()):
        del ws.tables[existing_table_name]

    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    ws.append([str(col) for col in df.columns.tolist()])
    for row in df.itertuples(index=False):
        ws.append(list(row))

    # Excel table objects with a header-only range are not consistently accepted
    # across readers. Keep the header row when there is no data, but skip creating
    # a table definition so the workbook remains valid.
    if len(df) == 0:
        return

    end_col = get_column_letter(len(df.columns))
    end_row = len(df) + 1
    tab = Table(displayName=table_name, ref=f"A1:{end_col}{end_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)




def _write_df_as_table_at(ws, df: pd.DataFrame, table_name: str, start_row: int):
    start_col = 1
    for col_idx, col in enumerate(df.columns.tolist(), start=start_col):
        ws.cell(row=start_row, column=col_idx, value=str(col))
    for ridx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for cidx, val in enumerate(list(row), start=start_col):
            ws.cell(row=ridx, column=cidx, value=val)

    if len(df) == 0:
        return start_row + 1

    end_col = get_column_letter(len(df.columns))
    end_row = start_row + len(df)
    tab = Table(displayName=table_name, ref=f"A{start_row}:{end_col}{end_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    return end_row + 2


def _safe_table_suffix(value: object) -> str:
    return re.sub(r"[^A-Za-z0-9_]", "_", str(value))


def build_input_workbook(include_sample_rows: bool = False) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    if include_sample_rows:
        tables = [
            (
                "FG_Input",
                "tblFG",
                pd.DataFrame(
                    [
                        {"FG Code": "FG_A", "Margin": 12.0},
                        {"FG Code": "FG_B", "Margin": 9.0},
                    ]
                ),
            ),
            (
                "BOM_Input",
                "tblBOM",
                pd.DataFrame(
                    [
                        {"FG Code": "FG_A", "RM Code": "RM_1", "QtyPerPair": 2.0},
                        {"FG Code": "FG_A", "RM Code": "RM_2", "QtyPerPair": 1.0},
                        {"FG Code": "FG_B", "RM Code": "RM_1", "QtyPerPair": 1.0},
                    ]
                ),
            ),
            (
                "FG_Caps",
                "tblFGPlanCap",
                pd.DataFrame(
                    [
                        {"FG Code": "FG_A", "Max Plan Qty": 120},
                        {"FG Code": "FG_B", "Max Plan Qty": 90},
                    ]
                ),
            ),
            (
                "RM_Availability",
                "tblRMAvail",
                pd.DataFrame(
                    [
                        {"RM Code": "RM_1", "Avail_Stock": 180, "Avail_StockPO": 240, "RM_Rate": 4.5},
                        {"RM Code": "RM_2", "Avail_Stock": 70, "Avail_StockPO": 110, "RM_Rate": 2.0},
                    ]
                ),
            ),
            (
                "Controls",
                "tblControl_2",
                pd.DataFrame(
                    [
                        {"Key": "Mode_Avail", "Value": "STOCK"},
                        {"Key": "Objective", "Value": "PLAN"},
                        {"Key": "PurchaseTargets", "Value": "25,50,75,100"},
                    ]
                ),
            ),
        ]
    else:
        tables = [
            (
                "FG_Input",
                "fg_master",
                pd.DataFrame([{"FG Code": "REPLACE_FG", "Margin": 1.0}]),
            ),
            (
                "BOM_Input",
                "bom_master",
                pd.DataFrame([{"FG Code": "REPLACE_FG", "RM Code": "REPLACE_RM", "QtyPerPair": 1.0}]),
            ),
            (
                "FG_Caps",
                "tblFGPlanCap",
                pd.DataFrame([{"FG Code": "REPLACE_FG", "Max Plan Qty": 1}]),
            ),
            (
                "RM_Availability",
                "tblRMAvail",
                pd.DataFrame([{"RM Code": "REPLACE_RM", "Avail_Stock": 1, "Avail_StockPO": 1, "RM_Rate": 1.0}]),
            ),
            (
                "Controls",
                "tblControl_2",
                pd.DataFrame(
                    [
                        {"Key": "Mode_Avail", "Value": "STOCK"},
                        {"Key": "Objective", "Value": "PLAN"},
                        {"Key": "PurchaseTargets", "Value": "25,50,75,100"},
                    ]
                ),
            ),
        ]

    for sheet_name, table_name, df in tables:
        ws = wb.create_sheet(sheet_name)
        _write_df_as_table(ws, df, table_name)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def write_output_excel(
    fg_df: pd.DataFrame,
    rm_df: pd.DataFrame,
    meta_df: pd.DataFrame,
    purchase_summary_df: pd.DataFrame,
    purchase_detail_df: pd.DataFrame,
    purchase_target_sheets: dict[object, dict[str, pd.DataFrame]] | None = None,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    ws_fg = wb.create_sheet("FG_Result")
    ws_rm = wb.create_sheet("RM_Diagnostic")
    ws_meta = wb.create_sheet("Run_Metadata")
    ws_purchase_summary = wb.create_sheet("Purchase_Summary")
    ws_purchase_detail = wb.create_sheet("Purchase_Detail")

    _write_df_as_table(ws_fg, fg_df, "tblFGResult")
    _write_df_as_table(ws_rm, rm_df, "tblRMDiagnostic")
    _write_df_as_table(ws_meta, meta_df, "tblRunMeta")
    _write_df_as_table(ws_purchase_summary, purchase_summary_df, "tblPurchaseSummary")
    _write_df_as_table(ws_purchase_detail, purchase_detail_df, "tblPurchaseDetail")

    purchase_target_sheets = purchase_target_sheets or {}
    target_keys = list(purchase_target_sheets.keys()) or [25, 50, 75, 100]
    for raw_key in target_keys:
        sheet_suffix = str(raw_key)
        ws = wb.create_sheet(f"Purchase_{sheet_suffix}")
        payload = purchase_target_sheets.get(raw_key, {})
        fg_target = payload.get("fg", pd.DataFrame(columns=["FG Code", "Cap", "TargetPairs", "OptQty", "Shortfall", "UnitMargin", "TotalMargin"]))
        rm_target = payload.get("rm", pd.DataFrame(columns=["RM Code", "AvailBase", "RM_Rate", "UsedAtPlan", "BuyQty", "BuyCost", "RemainingAfterBuy"]))
        key_suffix = _safe_table_suffix(raw_key)
        next_row = _write_df_as_table_at(ws, fg_target, f"tblPurchaseFG_{key_suffix}", 1)
        _write_df_as_table_at(ws, rm_target, f"tblPurchaseRM_{key_suffix}", next_row)

    final = BytesIO()
    wb.save(final)
    final.seek(0)
    return final.getvalue()
