from io import BytesIO

import openpyxl

from app.assets import DEMO_TABLES_DIR, load_demo_tables, sample_input_bytes, sample_output_bytes, workflow_preview_svg
from app.excel_io import diagnose_workbook_structure, load_tables_from_excel
from app.orchestrator import run_optimization
from app.validate import validate_inputs


def test_demo_table_sources_are_present():
    tables = load_demo_tables()
    assert (DEMO_TABLES_DIR / "fg_master.csv").exists()
    assert (DEMO_TABLES_DIR / "bom_master.csv").exists()
    assert (DEMO_TABLES_DIR / "tblFGPlanCap.csv").exists()
    assert (DEMO_TABLES_DIR / "tblRMAvail.csv").exists()
    assert (DEMO_TABLES_DIR / "tblControl_2.csv").exists()
    assert len(tables["fg_master"]) == 50
    assert len(tables["tblRMAvail"]) == 200


def test_bundled_sample_input_is_valid_and_runs_end_to_end():
    workbook_bytes = sample_input_bytes()
    diagnosis = diagnose_workbook_structure(workbook_bytes)
    assert diagnosis["issues"] == []

    tables = load_tables_from_excel(workbook_bytes)
    validate_inputs(tables)

    fg_df, rm_df, meta_df, purchase_summary_df, purchase_detail_df = run_optimization(
        tables,
        run_purchase_planner=True,
        purchase_target_fill_pcts="25,50,75,100",
    )

    assert len(fg_df) == 50
    assert len(rm_df) == 200
    assert len(meta_df) > 0
    assert len(purchase_summary_df) == 4
    assert not purchase_detail_df.empty


def test_bundled_sample_output_contains_expected_sheets_tables_and_preview_asset():
    workbook = openpyxl.load_workbook(BytesIO(sample_output_bytes()))
    preview_svg = workflow_preview_svg()
    assert "<svg" in preview_svg

    expected_sheetnames = {
        "FG_Result",
        "RM_Diagnostic",
        "Run_Metadata",
        "Purchase_Summary",
        "Purchase_Detail",
        "Purchase_25",
        "Purchase_50",
        "Purchase_75",
        "Purchase_100",
    }
    assert expected_sheetnames.issubset(workbook.sheetnames)

    assert "tblFGResult" in workbook["FG_Result"].tables
    assert "tblRMDiagnostic" in workbook["RM_Diagnostic"].tables
    assert "tblRunMeta" in workbook["Run_Metadata"].tables
    assert "tblPurchaseSummary" in workbook["Purchase_Summary"].tables
    assert "tblPurchaseDetail" in workbook["Purchase_Detail"].tables
