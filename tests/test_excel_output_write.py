from io import BytesIO

import openpyxl
import pandas as pd

from app.excel_io import _write_df_as_table, write_output_excel


def test_write_output_excel_writes_single_header_and_data_block_per_sheet():
    fg_df = pd.DataFrame(
        [
            {"FG Code": "FG1", "Units": 10},
            {"FG Code": "FG2", "Units": 20},
        ]
    )
    rm_df = pd.DataFrame([
        {"RM Code": "RM1", "Variance": 1.5},
    ])
    meta_df = pd.DataFrame([
        {"Metric": "run_id", "Value": "abc123"},
    ])

    output_bytes = write_output_excel(fg_df=fg_df, rm_df=rm_df, meta_df=meta_df)
    wb = openpyxl.load_workbook(BytesIO(output_bytes))

    expected = [
        ("FG_Result", fg_df, "tblFGResult"),
        ("RM_Diagnostic", rm_df, "tblRMDiagnostic"),
        ("Run_Metadata", meta_df, "tblRunMeta"),
    ]

    for sheet_name, df, table_name in expected:
        ws = wb[sheet_name]

        # Regression check: this used to be duplicated (2 * len(df) + 2 rows)
        # because the dataframe was written once by pandas and appended again.
        assert ws.max_row == len(df) + 1

        header_row = [cell.value for cell in ws[1]]
        assert header_row == df.columns.tolist()

        assert len(ws.tables) == 1
        table = next(iter(ws.tables.values()))
        assert table.displayName == table_name


def test_write_df_as_table_repeated_calls_replace_existing_table_and_cells():
    wb = openpyxl.Workbook()
    ws = wb.active

    first_df = pd.DataFrame([
        {"Code": "A", "Qty": 1},
        {"Code": "B", "Qty": 2},
    ])
    second_df = pd.DataFrame([
        {"Code": "X", "Qty": 9},
    ])

    _write_df_as_table(ws, first_df, "tblFirst")
    assert len(ws.tables) == 1
    assert ws.max_row == len(first_df) + 1

    _write_df_as_table(ws, second_df, "tblSecond")

    assert ws.max_row == len(second_df) + 1
    assert [cell.value for cell in ws[1]] == second_df.columns.tolist()
    assert [cell.value for cell in ws[2]] == ["X", 9]
    assert len(ws.tables) == 1
    assert "tblFirst" not in ws.tables
    assert "tblSecond" in ws.tables
    assert next(iter(ws.tables.values())).ref == "A1:B2"


def test_write_df_as_table_repeated_writes_persist_with_single_table_after_save_reload():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FG_Result"

    original_df = pd.DataFrame([
        {"FG Code": "FG1", "Units": 5},
        {"FG Code": "FG2", "Units": 8},
    ])
    rewritten_df = pd.DataFrame([
        {"FG Code": "FG9", "Units": 99},
    ])

    _write_df_as_table(ws, original_df, "tblFGResult")
    _write_df_as_table(ws, rewritten_df, "tblFGResult")

    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)
    reloaded = openpyxl.load_workbook(payload)
    out_ws = reloaded["FG_Result"]

    assert out_ws.max_row == len(rewritten_df) + 1
    assert [cell.value for cell in out_ws[1]] == rewritten_df.columns.tolist()
    assert [cell.value for cell in out_ws[2]] == ["FG9", 99]
    assert len(out_ws.tables) == 1
    table = next(iter(out_ws.tables.values()))
    assert table.displayName == "tblFGResult"
    assert table.ref == "A1:B2"
